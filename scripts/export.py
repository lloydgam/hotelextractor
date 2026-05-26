from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
EVEN_FILL   = PatternFill("solid", fgColor="D6E4F0")
COLUMNS     = ["#", "Hotel Name", "Price/Night (USD)", "Total Price (3n)", "Rating", "Location"]
COL_WIDTHS  = [5, 44, 20, 20, 10, 18]


def init_excel(path):
    """Create a fresh Excel workbook with a styled header row."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Hotels"
    for col, (header, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def append_hotel(hotel, row_num, path, lock):
    """Thread-safe: load workbook, append one hotel row, save."""
    row_data = [
        row_num,
        hotel["hotel_name"],
        hotel["price_usd"],
        hotel["total_price_usd"],
        hotel["rating"],
        hotel["location"],
    ]
    with lock:
        wb  = load_workbook(path)
        ws  = wb.active
        row = ws.max_row + 1
        fill = EVEN_FILL if row % 2 == 0 else None
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if fill:
                cell.fill = fill
        wb.save(path)


if __name__ == "__main__":
    import argparse, json, threading
    p = argparse.ArgumentParser(description="Exporter agent — writes hotels to Excel")
    p.add_argument("--hotels", required=True, help="JSON array of extracted hotel dicts (sorted)")
    p.add_argument("--path",   required=True, help="Output Excel file path")
    args = p.parse_args()

    hotels = json.loads(args.hotels)
    path   = Path(args.path)
    init_excel(path)
    lock   = threading.Lock()
    for i, hotel in enumerate(hotels, 1):
        append_hotel(hotel, i, path, lock)
    print(json.dumps({"status": "ok", "rows_written": len(hotels), "path": str(path)}))
